from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
import models, logic, json

def generate_product_card(db_produkt: models.Produkt, lang="pl"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Specyfikacja Produktu" if lang == "pl" else "Product Specification"
    
    # Styles
    header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    label_font = Font(name='Arial', size=10, bold=True)
    value_font = Font(name='Arial', size=10)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Header
    ws.merge_cells('A1:D1')
    ws['A1'] = "KARTA SPECYFIKACJI PRODUKTU" if lang == "pl" else "PRODUCT SPECIFICATION SHEET"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Basic Info
    ws['A3'] = "Nazwa produktu:" if lang == "pl" else "Product Name:"
    ws['B3'] = db_produkt.nazwa_pl if lang == "pl" else (db_produkt.nazwa_en or db_produkt.nazwa_pl)
    
    ws['A4'] = "Nazwa prawna produktu / opis produktu:" if lang == "pl" else "Legal Name / Description:"
    ws.merge_cells('B4:D4')
    ws['B4'] = db_produkt.prawna_nazwa_pl or "-"
    ws['B4'].alignment = Alignment(wrap_text=True)
    
    ws['A5'] = "Kod EAN / GTIN:"
    ws['B5'] = db_produkt.ean
    
    ws['A6'] = "Kod wewnętrzny:" if lang == "pl" else "Internal ID:"
    ws['B6'] = db_produkt.internal_id
    
    ws['A7'] = "Kod CN:"
    ws['B7'] = db_produkt.kod_cn or "-"
    
    ws['A8'] = "Kod PKWiU:"
    ws['B8'] = db_produkt.kod_pkwiu or "-"
    
    for r in range(3, 9):
        ws[f'A{r}'].font = label_font
        ws[f'B{r}'].font = value_font
        
    # Nutrition
    row = 10
    ws[f'A{row}'] = "WARTOŚCI ODŻYWCZE (100g)" if lang == "pl" else "NUTRITIONAL VALUES (100g)"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    
    nutrition = logic.calculate_nutrition(db_produkt)
    row = 8
    labels = {
        "energia_kj": "Energia Kj:",
        "energia_kcal": "Energia Kcal:",
        "tluszcz": "Tluszcz:",
        "kwasy_nasycone": "Kwasy Nasycone:",
        "weglowodany": "Weglowodany:",
        "cukry": "Cukry:",
        "bialko": "Bialko:",
        "sol": "Sól:",
        "blonnik": "Błonnik:"
    } if lang == "pl" else {
        "energia_kj": "Energy Kj:",
        "energia_kcal": "Energy Kcal:",
        "tluszcz": "Fat:",
        "kwasy_nasycone": "  of which saturates:",
        "weglowodany": "Carbohydrate:",
        "cukry": "  of which sugars:",
        "bialko": "Protein:",
        "sol": "Salt:",
        "blonnik": "Fiber:"
    }
    
    for key, label in labels.items():
        ws[f'A{row}'] = label
        ws[f'B{row}'] = round(nutrition.get(key, 0), 1)
        ws[f'A{row}'].font = Font(size=9)
        ws[f'B{row}'].font = Font(size=9)
        row += 1
        
    # Ingredients
    ws.merge_cells(f'A{row+1}:D{row+1}')
    ws[f'A{row+1}'] = "SKŁAD" if lang == "pl" else "INGREDIENTS"
    ws[f'A{row+1}'].font = Font(bold=True, size=11)
    
    sklad_text = logic.generate_ingredients_text(db_produkt, lang=lang)
    ws.merge_cells(f'A{row+2}:D{row+4}')
    ws[f'A{row+2}'] = sklad_text
    ws[f'A{row+2}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws[f'A{row+2}'].font = Font(size=9)

    # Ingredient Origins Section
    row = row + 6
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "PROCENTOWY UDZIAŁ SKŁADNIKÓW I KRAJE POCHODZENIA" if lang == "pl" else "INGREDIENT PERCENTAGE AND COUNTRIES OF ORIGIN"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    
    row += 1
    # Adding headers for the table
    ws[f'A{row}'] = "Składnik" if lang == "pl" else "Ingredient"
    ws[f'B{row}'] = "Udział %" if lang == "pl" else "Share %"
    ws[f'C{row}'] = "Kraje pochodzenia" if lang == "pl" else "Countries of origin"
    ws[f'A{row}'].font = Font(bold=True, size=9)
    ws[f'B{row}'].font = Font(bold=True, size=9)
    ws[f'C{row}'].font = Font(bold=True, size=9)
    row += 1

    ingredient_origins = logic.get_ingredient_origins(db_produkt)
    for item in ingredient_origins:
        ws[f'A{row}'] = item['name']
        ws[f'B{row}'] = f"{str(round(item['percent'], 2)).replace('.', ',')}%"
        ws[f'C{row}'] = ", ".join(item['countries']) if item['countries'] else "-"
        
        ws[f'A{row}'].font = Font(size=9)
        ws[f'B{row}'].font = Font(size=9)
        ws[f'C{row}'].font = Font(size=9)
        row += 1
    
    # Organoleptyka Section
    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "ORGANOLEPTYKA" if lang == "pl" else "ORGANOLEPTIC PROPERTIES"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    
    row += 1
    ws[f'A{row}'] = "Parametr" if lang == "pl" else "Parameter"
    ws[f'B{row}'] = "Opis" if lang == "pl" else "Description"
    ws[f'A{row}'].font = Font(bold=True, size=9)
    ws[f'B{row}'].font = Font(bold=True, size=9)
    row += 1
    
    organo_fields = [
        ("SMAK" if lang == "pl" else "TASTE", db_produkt.organoleptyka_smak),
        ("ZAPACH" if lang == "pl" else "SMELL", db_produkt.organoleptyka_zapach),
        ("KOLOR" if lang == "pl" else "COLOUR", db_produkt.organoleptyka_kolor),
        ("WYGLĄD ZEWNĘTRZNY" if lang == "pl" else "EXTERNAL APPEARANCE", db_produkt.organoleptyka_wyglad_zewnetrzny),
        ("WYGLĄD NA PRZEKROJU" if lang == "pl" else "CROSS-SECTION VIEW", db_produkt.organoleptyka_wyglad_na_przekroju)
    ]
    
    for label, value in organo_fields:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value or "-"
        ws[f'A{row}'].font = Font(size=9)
        ws[f'B{row}'].font = Font(size=9)
        ws[f'B{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    
    # Storage, Shelf Life and Additional Information
    row += 1
    new_sections = [
        ("WARUNKI PRZECHOWYWANIA I TRANSPORTU" if lang == "pl" else "STORAGE AND TRANSPORT CONDITIONS", db_produkt.warunki_przechowywania),
        ("TERMIN PRZYDATNOŚCI DO SPOŻYCIA" if lang == "pl" else "SHELF LIFE", db_produkt.termin_przydatnosci),
        ("WYRAŻENIE I FORMAT" if lang == "pl" else "DATE EXPRESSION AND FORMAT", db_produkt.wyrazenie_format_daty),
        ("INFORMACJE DODATKOWE" if lang == "pl" else "ADDITIONAL INFORMATION", db_produkt.informacje_dodatkowe)
    ]
    
    for label, value in new_sections:
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True, size=11)
        row += 1
        ws.merge_cells(f'A{row}:D{row+1}')
        ws[f'A{row}'] = value or "-"
        ws[f'A{row}'].font = Font(size=9)
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        row += 3
    
    # LOGISTYKA Section
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "LOGISTYKA" if lang == "pl" else "LOGISTICS"
    ws[f'A{row}'].font = Font(bold=True, size=11, color='1F4E78')
    row += 1

    # Wymiary Table
    ws[f'A{row}'] = "1. WYMIARY [cm]" if lang == "pl" else "1. DIMENSIONS [cm]"
    ws[f'A{row}'].font = Font(bold=True, size=10)
    row += 1
    
    headers = ["Poziom", "Wysokość", "Szerokość", "Głębokość"] if lang == "pl" else ["Level", "Height", "Width", "Depth"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=i+1)
        cell.value = h
        cell.font = Font(bold=True, size=9)
    row += 1
    
    dim_rows = [
        ("Produkt solo", "logistyka_wymiary_solo"),
        ("W opakowaniu jednostkowym", "logistyka_wymiary_jednostka"),
        ("Opakowanie zbiorcze 1°", "logistyka_wymiary_zbiorcze1"),
        ("Opakowanie zbiorcze 2°", "logistyka_wymiary_zbiorcze2"),
        ("Opakowanie zbiorcze 3°", "logistyka_wymiary_zbiorcze3"),
    ] if lang == "pl" else [
        ("Product solo", "logistyka_wymiary_solo"),
        ("Unit packaging", "logistyka_wymiary_jednostka"),
        ("Master case 1°", "logistyka_wymiary_zbiorcze1"),
        ("Master case 2°", "logistyka_wymiary_zbiorcze2"),
        ("Master case 3°", "logistyka_wymiary_zbiorcze3"),
    ]
    
    for label, prefix in dim_rows:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = getattr(db_produkt, f"{prefix}_h", 0)
        ws[f'C{row}'] = getattr(db_produkt, f"{prefix}_w", 0)
        ws[f'D{row}'] = getattr(db_produkt, f"{prefix}_d", 0)
        for i in range(1, 5):
            ws.cell(row=row, column=i).font = Font(size=9)
        row += 1
    
    ws[f'A{row}'] = "Rodzaj palety:" if lang == "pl" else "Pallet type:"
    ws[f'B{row}'] = db_produkt.logistyka_rodzaj_palety or "-"
    ws[f'A{row}'].font = Font(bold=True, size=9)
    ws[f'B{row}'].font = Font(size=9)
    row += 2

    # Wagi Table
    ws[f'A{row}'] = "2. WAGI [kg]" if lang == "pl" else "2. WEIGHTS [kg]"
    ws[f'A{row}'].font = Font(bold=True, size=10)
    row += 1
    
    weight_rows = [
        ("Waga netto sztuki", "logistyka_waga_netto_szt"),
        ("Waga brutto sztuki", "logistyka_waga_brutto_szt"),
        ("Waga netto opakowania zbiorczego", "logistyka_waga_netto_zbiorcze"),
        ("Waga brutto opakowania zbiorczego", "logistyka_waga_brutto_zbiorcze"),
        ("Waga netto palety", "logistyka_waga_netto_paleta"),
        ("Waga brutto palety", "logistyka_waga_brutto_paleta"),
    ] if lang == "pl" else [
        ("Unit net weight", "logistyka_waga_netto_szt"),
        ("Unit gross weight", "logistyka_waga_brutto_szt"),
        ("Master case net weight", "logistyka_waga_netto_zbiorcze"),
        ("Master case gross weight", "logistyka_waga_brutto_zbiorcze"),
        ("Pallet net weight", "logistyka_waga_netto_paleta"),
        ("Pallet gross weight", "logistyka_waga_brutto_paleta"),
    ]
    
    for label, field in weight_rows:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = getattr(db_produkt, field, 0)
        ws[f'A{row}'].font = Font(size=9)
        ws[f'B{row}'].font = Font(size=9)
        row += 1
    row += 1

    # Paletyzacja Table
    ws[f'A{row}'] = "3. PALETYZACJA" if lang == "pl" else "3. PALLETIZATION"
    ws[f'A{row}'].font = Font(bold=True, size=10)
    row += 1
    
    pallet_rows = [
        ("Ilość sztuk w opakowaniu zbiorczym", "logistyka_sztuk_w_zbiorczym", "szt"),
        ("Ilość kartonów na warstwie", "logistyka_kartonow_na_warstwie", "szt"),
        ("Ilość warstw na palecie", "logistyka_warstw_na_palecie", "szt"),
        ("Ilość kartonów na palecie", "logistyka_kartonow_na_palecie", "szt"),
        ("Ilość sztuk na palecie", "logistyka_sztuk_na_palecie", "szt"),
        ("Ilość sztuk na warstwie palety", "logistyka_sztuk_na_warstwie", "szt"),
        ("Wysokość palety z nośnikiem", "logistyka_wysokosc_palety", "cm"),
    ] if lang == "pl" else [
        ("Quantity per master case", "logistyka_sztuk_w_zbiorczym", "pcs"),
        ("Cases per layer", "logistyka_kartonow_na_warstwie", "pcs"),
        ("Layers per pallet", "logistyka_warstw_na_palecie", "pcs"),
        ("Cases per pallet", "logistyka_kartonow_na_palecie", "pcs"),
        ("Quantity per pallet", "logistyka_sztuk_na_palecie", "pcs"),
        ("Quantity per layer", "logistyka_sztuk_na_warstwie", "pcs"),
        ("Pallet height (incl. carrier)", "logistyka_wysokosc_palety", "cm"),
    ]
    
    for label, field, unit in pallet_rows:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = getattr(db_produkt, field, 0)
        ws[f'C{row}'] = unit
        ws[f'A{row}'].font = Font(size=9)
        ws[f'B{row}'].font = Font(size=9)
        ws[f'C{row}'].font = Font(size=9)
        row += 1
        
    # Certificates section
    import json
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "CERTYFIKATY" if lang == "pl" else "CERTIFICATES"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1
    
    ws[f'A{row}'] = "Rodzaj" if lang == "pl" else "Type"
    ws[f'B{row}'] = "COID"
    ws[f'C{row}'] = "Data ważności certyfikatu" if lang == "pl" else "Expiry Date"
    ws[f'A{row}'].font = Font(bold=True, size=9)
    ws[f'B{row}'].font = Font(bold=True, size=9)
    ws[f'C{row}'].font = Font(bold=True, size=9)
    row += 1
    
    try:
        certs = json.loads(db_produkt.certyfikaty or "[]")
        for cert in certs:
            ws[f'A{row}'] = cert.get('rodzaj', '-')
            ws[f'B{row}'] = cert.get('coid', '-')
            ws[f'C{row}'] = cert.get('data_waznosci', '-')
            ws[f'A{row}'].font = Font(size=9)
            ws[f'B{row}'].font = Font(size=9)
            ws[f'C{row}'].font = Font(size=9)
            row += 1
    except:
        ws[f'A{row}'] = "Brak danych" if lang == "pl" else "No data"
        row += 1
    
    # Allergens
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "ALERGENY" if lang == "pl" else "ALLERGENS"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    
    allergens = logic.aggregate_allergens(db_produkt)
        
    allergen_ids = [
        "gluten", "skorupiaki", "jaja", "ryby", "orzeszki_ziemne", "soja", 
        "mleko", "orzechy", "seler", "gorczyca", "sezam", "dwutlenek_siarki", 
        "lubin", "mieczaki"
    ]
    
    allergen_map = {
        'gluten': {'pl': 'Gluten', 'en': 'Gluten'},
        'skorupiaki': {'pl': 'Skorupiaki', 'en': 'Crustaceans'},
        'jaja': {'pl': 'Jaja', 'en': 'Eggs'},
        'ryby': {'pl': 'Ryby', 'en': 'Fish'},
        'orzeszki_ziemne': {'pl': 'Orzeszki ziemne', 'en': 'Peanuts'},
        'soja': {'pl': 'Soja', 'en': 'Soybeans'},
        'mleko': {'pl': 'Mleko', 'en': 'Milk'},
        'orzechy': {'pl': 'Orzechy', 'en': 'Nuts'},
        'seler': {'pl': 'Seler', 'en': 'Celery'},
        'gorczyca': {'pl': 'Gorczyca', 'en': 'Mustard'},
        'sezam': {'pl': 'Sezam', 'en': 'Sesame seeds'},
        'dwutlenek_siarki': {'pl': 'Dwutlenek siarki i siarczyny', 'en': 'Sulphur dioxide and sulphites'},
        'lubin': {'pl': 'Łubin', 'en': 'Lupin'},
        'mieczaki': {'pl': 'Mięczaki', 'en': 'Molluscs'},
    }

    status_map = {
        'Nie zawiera': {'pl': 'Nie zawiera', 'en': 'Does not contain'},
        'Zawiera': {'pl': 'Zawiera', 'en': 'Contains'},
        'Może zawierać': {'pl': 'Może zawierać', 'en': 'May contain'},
    }
    
    row += 1
    for alg_id in allergen_ids:
        status = allergens.get(alg_id, "Nie zawiera")
        name = allergen_map.get(alg_id, {}).get(lang, alg_id.replace('_', ' ').capitalize())
        status_label = status_map.get(status, {}).get(lang, status)
        
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = f"{name} – {status_label}"
        ws[f'A{row}'].font = Font(size=9)
        
        if status == "Zawiera":
             ws[f'A{row}'].font = Font(size=9, color='C00000', bold=True)
        elif status == "Może zawierać":
             ws[f'A{row}'].font = Font(size=9, color='FFC000')
             
        row += 1
    
    # Auto-adjust columns
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
